import requests
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import os

headers = {
    "X-QuakeToken": "your Quake_Token"
}


def create_excel_file(file_path, sheet_name, headers):
    workbook = Workbook()
    if os.path.exists(file_path):  # 如果文件存在，则删除
        os.remove(file_path)
    workbook.remove(workbook.active)  # 移除默认工作表
    workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
    ws = workbook[sheet_name]
    ws.append(headers)  # 写入表头
    workbook.save(file_path)


def for_ip_cn():

    file_path = 'for_ip_cn.xlsx'  # Excel 文件路径
    sheet_name = 'Sheet1'  # 工作表名称
    headers = ['IP', '证书']  # 表头内容
    if not os.path.exists(file_path):  # 如果文件不存在，则创建
        create_excel_file(file_path, sheet_name, headers)
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)  # 移除默认工作表
        workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
        ws = workbook[sheet_name]
        ws.append(headers)  # 写入表头
    else:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
            ws = workbook[sheet_name]
            ws.append(headers)  # 写入表头
        else:
            ws = workbook[sheet_name]

    data = datas['data']
    result_list = []
    for i in data:
        # print("IP：" + i['ip'])  # 获取IP
        # print(i)
        service = i['service']
        try:    # 获取证书域名
            tls = service['tls']
            handshake_log = tls['handshake_log']
            server_certificates = handshake_log['server_certificates']
            certificate = server_certificates['certificate']
            parsed = certificate['parsed']

            cn = parsed['subject_dn']
            ws.append([i['ip'], cn])
            result_list.append([i['ip'], cn])

        except:
            pass
    workbook.save(file_path)  # 保存文件
    workbook.close()
    return result_list


def for_domain():

    file_path = 'for_ip_domain.xlsx'  # Excel 文件路径
    sheet_name = 'Sheet1'  # 工作表名称
    headers = ['IP', '域名']  # 表头内容
    if not os.path.exists(file_path):  # 如果文件不存在，则创建
        create_excel_file(file_path, sheet_name, headers)
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)  # 移除默认工作表
        workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
        ws = workbook[sheet_name]
        ws.append(headers)  # 写入表头
    else:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
            ws = workbook[sheet_name]
            ws.append(headers)  # 写入表头
        else:
            ws = workbook[sheet_name]
    # 获取url
    data = datas['data']
    resulte_set = set()
    result_list = []
    for i in data:
        try:
            httpss = i['service']['http']
            ip_domain = httpss['http_load_url']
            # print(ip_domain)
            for item in ip_domain:
                domain_match = re.search(r'(?<=://)([^:/]+)', item)
                if domain_match:
                    domain = domain_match.group(0)
                    is_ip_address = re.match(r'\d+\.\d+\.\d+\.\d+', domain)
                    if not is_ip_address:
                        # print(domain)
                        resulte_data = [i['ip'], domain]
                        if tuple(resulte_data) not in resulte_set:

                            ws.append([i['ip'], domain])
                            result_list.append([i['ip'], domain])
                            resulte_set.add(tuple(resulte_data))
        except:
            pass

    workbook.save(file_path)  # 保存文件
    workbook.close()
    return result_list


def for_ip_zic():

    file_path = 'for_ip_zic.xlsx'  # Excel 文件路径
    sheet_name = 'Sheet1'  # 工作表名称
    headers = ['ip', '资产类型', '资产名称']  # 表头内容
    if not os.path.exists(file_path):  # 如果文件不存在，则创建
        create_excel_file(file_path, sheet_name, headers)
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.remove(workbook.active)  # 移除默认工作表
        workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
        ws = workbook[sheet_name]
        ws.append(headers)  # 写入表头
    else:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name, index=0)  # 创建新的工作表
            ws = workbook[sheet_name]
            ws.append(headers)  # 写入表头
        else:
            ws = workbook[sheet_name]
    # 获取url
    data = datas['data']
    resulte_set = set()
    result_list = []
    for i in data:
        try:
            httpss = i['service']['http']
            ip_domain = httpss['http_load_url']
            components = i['components']
            scan = components[0]['product_type'][0]
            product_name_cn = components[0]['product_name_cn']
            # print(product_name_cn)
            if scan == "安全扫描与检测(Scanner)":
                for j in ip_domain:
                    # print(j)
                    print(scan, product_name_cn, j)
                    ws.append([i['ip'], scan, product_name_cn, j])
                    result_list.append([i['ip'], scan, product_name_cn, j])
        except:
            pass

    workbook.save(file_path)  # 保存文件
    workbook.close()
    return result_list


if __name__ == '__main__':
    if os.path.exists('for_ip_cn.xlsx'):  # 如果文件存在，则删除
        os.remove('for_ip_cn.xlsx')
    if os.path.exists('for_ip_domain.xlsx'):  # 如果文件存在，则删除
        os.remove('for_ip_domain.xlsx')
    if os.path.exists('for_ip_zic.xlsx'):  # 如果文件存在，则删除
        os.remove('for_ip_zic.xlsx')
    file_path = 'ip_addresses.txt'  # 文本文件路径
    with open(file_path, 'r') as file:
        ip_addresses = file.readlines()
    for ipd in ip_addresses:
        ips = ipd.strip()  # 去除换行符和空格
        print(ips)
        # ips = '119.27.174.182'

        data = {
            "query": "ip:" + ips,
            "start": 0,
            "size": 100
        }

        response = requests.post(url="https://quake.360.net/api/v3/search/quake_service", headers=headers, json=data)
        try:
            datas = response.json()
        except:
            pass
        for_ip_cn()
        for_domain()
        for_ip_zic()



